#!/usr/bin/env python3
"""
Montador de relatório — PUSH DOE-SC / DETRAN.

Lê a saída do motor (resultados.json em CSV_DIR; ou, na falta dele, os CSVs por
edição) e produz, na pasta de saída:
  - Buscador_Publicacoes_DOE.xlsx  — planilha de controle com semáforo de prazos
    e um bloco de controle (amarelo) preservado entre execuções pelo ID da matéria.
  - Dashboard_DOE.html             — painel autossuficiente (KPIs, barras e tabela
    filtrável), abre em qualquer navegador sem internet.
  - Versões/<arquivo>_AAAA-MM-DD.(xlsx|html) — cópia datada de cada execução.

Ao final imprime uma linha JSON com o resumo.

Uso:
  python3 montar_relatorio.py --entrada "<pasta>/resultados.json" --saida "<pasta>"
  python3 montar_relatorio.py --csv-dir "<pasta com CSVs>" --saida "<pasta>"
Opções:
  --urgente N   dias corridos até a data final para marcar urgência/vermelho (padrão 5)
  --titulo T    rótulo opcional no topo dos arquivos
Dependências: openpyxl
"""
import argparse, json, os, csv, re, html, datetime, shutil, collections, sys, glob

HOJE = datetime.date.today()
EDIT = ["Status", "Responsável", "Prazo fatal (confirmado)", "Providência", "Data da ciência"]
# Marca em ordem cronológica de vencimento; sem prazo ao final.
ORD = {"venc": 0, "r": 1, "y": 2, "g": 3, "sem": 4}
COR_HEX = {"venc": "7b241c", "r": "e74c3c", "y": "e0a800", "g": "28a745", "sem": "bbbbbb"}


def parse_args():
    ap = argparse.ArgumentParser()
    ap.add_argument("--entrada", default="", help="caminho do resultados.json")
    ap.add_argument("--csv-dir", default="", help="pasta com os CSVs por edição (fallback)")
    ap.add_argument("--saida", default="", help="pasta de saída (padrão: a do resultados.json/csv-dir)")
    ap.add_argument("--urgente", type=int, default=5)
    ap.add_argument("--titulo", default="")
    return ap.parse_args()


def parse_data(s):
    if not s:
        return None
    s = str(s).strip()
    for fmt in ("%d/%m/%Y", "%Y-%m-%d"):
        try:
            return datetime.datetime.strptime(s, fmt).date()
        except Exception:
            pass
    return None


def estado(dfinal):
    """Classifica a situação do prazo por dias corridos até a data final."""
    if not dfinal:
        return "sem", None
    dias = (dfinal - HOJE).days
    if dias < 0:
        return "venc", dias
    if dias <= 0:  # vence hoje
        return "r", dias
    return ("r" if dias <= ARGS_URGENTE else "y" if dias <= 10 else "g"), dias


def carregar_registros(args):
    """Devolve lista de dicts normalizados a partir do resultados.json ou dos CSVs."""
    entrada = args.entrada
    if not entrada and args.csv_dir:
        cand = os.path.join(args.csv_dir, "resultados.json")
        if os.path.exists(cand):
            entrada = cand
    regs = []
    if entrada and os.path.exists(entrada):
        data = json.load(open(entrada, encoding="utf-8"))
        for r in data.get("resultados", []):
            regs.append({
                "id": str(r.get("id", "")),
                "edicao": str(r.get("edicao", "")),
                "statusBusca": r.get("status", ""),
                "motivo": r.get("motivoConferencia") or "",
                "cliente": r.get("cliente", ""),
                "cnh": r.get("cnh", ""),
                "processo": r.get("processo", ""),
                "tipo": r.get("tipoDecisao", ""),
                "dataPub": r.get("dataPublicacao", ""),
                "prazo": r.get("prazo"),
                "dataFinal": r.get("dataFinal") or "",
                "link": r.get("link", ""),
            })
        return regs, entrada
    # Fallback: ler os CSVs por edição (sem ID — chave de controle por edição+cliente+processo)
    csv_dir = args.csv_dir or (os.path.dirname(entrada) if entrada else ".")
    for fp in glob.glob(os.path.join(csv_dir, "*.csv")):
        with open(fp, encoding="utf-8-sig") as f:
            linhas = list(csv.reader(f, delimiter=";"))
        edicao = ""
        if linhas and linhas[0]:
            m = re.search(r"Edição Nº\s*([\w-]+)", linhas[0][0])
            edicao = m.group(1) if m else ""
        try:
            h = next(i for i, l in enumerate(linhas) if l and l[0] == "Status")
        except StopIteration:
            continue
        for l in linhas[h + 1:]:
            if len(l) < 8 or not l[1].strip():
                continue
            st = l[0]
            regs.append({
                "id": f"{edicao}|{l[1]}|{l[3]}",
                "edicao": edicao,
                "statusBusca": "CONFIRMADO" if st.startswith("CONFIRMADO") else "A CONFERIR",
                "motivo": "",
                "cliente": l[1], "cnh": l[2], "processo": l[3], "tipo": l[4],
                "dataPub": l[5], "prazo": l[6], "dataFinal": l[7], "link": "",
            })
    return regs, (entrada or csv_dir)


def ler_controle_anterior(xlsx_path):
    """Lê o bloco de controle da planilha anterior, indexado por ID."""
    prev = {}
    if not os.path.exists(xlsx_path):
        return prev
    try:
        from openpyxl import load_workbook
        wb = load_workbook(xlsx_path)
        ws = wb["Publicações"]
        hdr = [c.value for c in ws[1]]
        if "ID" not in hdr:
            return prev
        ci = hdr.index("ID")
        for r in range(2, ws.max_row + 1):
            vals = [c.value for c in ws[r]]
            idv = vals[ci] if ci < len(vals) else None
            if idv is None:
                continue
            prev[str(idv)] = {col: (vals[hdr.index(col)] if col in hdr and hdr.index(col) < len(vals) else None) for col in EDIT}
    except Exception as e:
        print(f"[aviso] não foi possível ler o controle anterior ({e}); o bloco pode reiniciar.", file=sys.stderr)
    return prev


def gerar_xlsx(regs, saida, titulo):
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    from openpyxl.utils import get_column_letter

    ATUAL = os.path.join(saida, "Buscador_Publicacoes_DOE.xlsx")
    prev = ler_controle_anterior(ATUAL)
    for r in regs:
        pe = prev.get(r["id"], {})
        r["ctrl"] = {c: (pe.get(c) or "") for c in EDIT}
        if not r["ctrl"]["Status"]:
            r["ctrl"]["Status"] = "Novo"

    cols = [
        "Status busca", "Cliente", "CNH", "Nº Processo administrativo", "Tipo de decisão",
        "Data publicação", "Prazo (dias)", "Data final p/ recurso",
    ] + EDIT + ["Edição", "ID", "Link oficial"]

    TIT = Font(name="Poppins", size=11, bold=True, color="000000")
    HDR = Font(name="Poppins", size=11, bold=True, color="000000")
    BODY = Font(name="Lora", size=11, color="000000")
    LINKF = Font(name="Lora", size=11, color="0000EE", underline="single")
    thin = Side(style="thin", color="D9D9D9")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    GRAY = PatternFill("solid", fgColor="F2F2F2")
    YEL = PatternFill("solid", fgColor="FFF6CC")
    SEMA = {
        "venc": PatternFill("solid", fgColor="F2B8B5"),
        "r": PatternFill("solid", fgColor="F9D5D3"),
        "y": PatternFill("solid", fgColor="FFEAB0"),
        "g": PatternFill("solid", fgColor="D4EDDA"),
        "sem": None,
    }

    wb = Workbook()
    ws = wb.active
    ws.title = "Publicações"
    ws.append(cols)
    for c in range(1, len(cols) + 1):
        cell = ws.cell(1, c)
        cell.font = HDR
        cell.alignment = Alignment(vertical="center", horizontal="center", wrap_text=True)
        cell.border = border

    lc = cols.index("Link oficial") + 1
    edit_idx = {cols.index(c) + 1 for c in EDIT}
    cnt = collections.Counter()
    for i, r in enumerate(regs, start=2):
        cnt[r["_state"]] += 1
        valores = [
            r["statusBusca"] + (f" ({r['motivo']})" if r["motivo"] and r["statusBusca"] != "CONFIRMADO" else ""),
            r["cliente"], r["cnh"], r["processo"], r["tipo"],
            r["dataPub"], r["prazo"] if r["prazo"] is not None else "", r["dataFinal"],
            r["ctrl"]["Status"], r["ctrl"]["Responsável"], r["ctrl"]["Prazo fatal (confirmado)"],
            r["ctrl"]["Providência"], r["ctrl"]["Data da ciência"],
            r["edicao"], r["id"], r["link"],
        ]
        ws.append(valores)
        baseF = SEMA[r["_state"]] or (GRAY if i % 2 == 0 else None)
        for c in range(1, len(cols) + 1):
            cell = ws.cell(i, c)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = border
            cell.font = BODY
            if baseF:
                cell.fill = baseF
            if c in edit_idx:
                cell.fill = YEL
            if c == lc and isinstance(cell.value, str) and cell.value[:4].lower() == "http":
                cell.hyperlink = cell.value
                cell.value = "abrir publicação"
                cell.font = LINKF

    W = {
        "Status busca": 22, "Cliente": 30, "CNH": 14, "Nº Processo administrativo": 22,
        "Tipo de decisão": 22, "Data publicação": 14, "Prazo (dias)": 11, "Data final p/ recurso": 17,
        "Status": 13, "Responsável": 16, "Prazo fatal (confirmado)": 17, "Providência": 22,
        "Data da ciência": 14, "Edição": 10, "ID": 14, "Link oficial": 16,
    }
    for idx, c in enumerate(cols, 1):
        ws.column_dimensions[get_column_letter(idx)].width = W.get(c, 14)
    ws.freeze_panes = "B2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(cols))}{len(regs) + 1}"

    obs = wb.create_sheet("Observações")
    conf = sum(1 for r in regs if r["statusBusca"] == "CONFIRMADO")
    linhas_obs = [
        titulo or "PUSH DOE-SC / DETRAN — Controle de publicações",
        f"Gerado em: {HOJE.strftime('%d/%m/%Y')}",
        f"Total: {len(regs)} · Confirmados: {conf} · A conferir: {len(regs) - conf}.",
        f"Situação de prazo — vencidos(est.): {cnt['venc']} · ≤{ARGS_URGENTE} dias: {cnt['r']} · 6–10 dias: {cnt['y']} · 11+ dias: {cnt['g']} · sem prazo: {cnt['sem']}.",
        "",
        f"Semáforo por dias corridos até a data final de recurso: vermelho escuro = vencido (estimado); vermelho ≤{ARGS_URGENTE}; amarelo 6–10; verde 11+; sem cor = sem prazo no edital.",
        "Ordenação: cronológica crescente (vence primeiro no topo); sem prazo ao final.",
        "CONFIRMADO = nome do cliente e CNH cadastrada bateram com o edital. A CONFERIR = nome bateu, mas o edital não trouxe CNH (ou o cliente está sem CNH cadastrada) — exige checagem manual.",
        "Bloco de controle (amarelo), preenchido pela controladoria e preservado por ID entre execuções: Status, Responsável, Prazo fatal (confirmado), Providência, Data da ciência. Fluxo sugerido de Status: Novo → Em análise → Em providência → Cumprido.",
        "Prazos são apoio: a data final é estimada a partir do prazo escrito no edital e da data de publicação. A data fatal real é a conferida nos autos pela controladoria.",
        "Dados pessoais (nome/CNH): uso interno do escritório (LGPD). Versionamento: cópia datada em 'Versões/'.",
    ]
    for i, l in enumerate(linhas_obs, 1):
        obs.cell(i, 1, l).font = Font(name=("Poppins" if i in (1, 6) else "Lora"), size=11, bold=(i in (1, 6)), color="000000")
    obs.column_dimensions["A"].width = 120

    wb.save(ATUAL)
    versoes = os.path.join(saida, "Versões")
    os.makedirs(versoes, exist_ok=True)
    shutil.copy2(ATUAL, os.path.join(versoes, f"Buscador_Publicacoes_DOE_{HOJE.strftime('%Y-%m-%d')}.xlsx"))
    return ATUAL, cnt


def gerar_html(regs, saida, titulo, cnt):
    esc = lambda s: html.escape(str(s if s is not None else ""))
    DASH = os.path.join(saida, "Dashboard_DOE.html")
    total = len(regs)
    conf = sum(1 for r in regs if r["statusBusca"] == "CONFIRMADO")
    aconf = total - conf
    por_tipo = dict(collections.Counter((r["tipo"] or "—") for r in regs).most_common())

    def bars(d, color="#5b8def"):
        if not d:
            return "<div class=muted>—</div>"
        mx = max(d.values())
        return "".join(
            f"<div class=bar><span class=bl>{esc(k)}</span><span class=bt style='width:{max(6,int(v/mx*100))}%;background:{color}'></span><span class=bn>{v}</span></div>"
            for k, v in d.items()
        )

    def bars_estado():
        lab = {"venc": "Vencido (est.)", "r": f"≤{ARGS_URGENTE} dias", "y": "6–10 dias", "g": "11+ dias", "sem": "Sem prazo"}
        mx = max(cnt.values()) if cnt else 1
        return "".join(
            f"<div class=bar><span class=bl>{lab[k]}</span><span class=bt style='width:{max(6,int(cnt.get(k,0)/mx*100))}%;background:#{COR_HEX[k]}'></span><span class=bn>{cnt.get(k,0)}</span></div>"
            for k in ["venc", "r", "y", "g", "sem"]
        )

    def trrow(r):
        df = parse_data(r["dataFinal"])
        venc = f"<span class=pill style='background:#{COR_HEX[r['_state']]}'>{esc(r['dataFinal'])}</span>" if df else "<span class=muted>—</span>"
        dias = "—" if r["_dias"] is None else (f"venc {abs(r['_dias'])}" if r["_dias"] < 0 else r["_dias"])
        lk = f"<a href='{esc(r['link'])}' target=_blank rel=noopener>abrir</a>" if (r["link"][:4].lower() == "http") else ""
        txt = esc((r["cliente"] + " " + r["processo"] + " " + r["tipo"] + " " + r["cnh"]).lower())
        return (
            f"<tr class={r['_state']} data-sb=\"{esc(r['statusBusca'])}\" data-tipo=\"{esc(r['tipo'])}\" data-state=\"{r['_state']}\" data-txt=\"{txt}\">"
            f"<td>{venc}</td><td>{dias}</td><td>{esc(r['dataPub'])}</td><td>{esc(r['cliente'])}</td><td>{esc(r['cnh'])}</td>"
            f"<td>{esc(r['processo'])}</td><td>{esc(r['tipo'])}</td><td>{esc(r['statusBusca'])}</td><td>{esc(r['ctrl']['Status'])}</td><td>{lk}</td></tr>"
        )

    TBODY = "".join(trrow(r) for r in regs)
    tipos = sorted(set(r["tipo"] for r in regs if r["tipo"]))
    opt = lambda L: "".join(f"<option value=\"{esc(x)}\">{esc(x)}</option>" for x in L)
    sub = (titulo + " · ") if titulo else ""

    HTML = f"""<!doctype html><html lang=pt-br><head><meta charset=utf-8><meta name=viewport content="width=device-width,initial-scale=1"><title>Painel de Publicações — DOE-SC / DETRAN</title>
<style>
@import url('https://fonts.googleapis.com/css2?family=Poppins:wght@500;600;700&family=Lora:wght@400;500&display=swap');
:root{{color-scheme:light}}*{{box-sizing:border-box}}
body{{font-family:'Lora',Georgia,serif;margin:0;background:#fafafa;color:#000}}
h1,h3,.kpi .n,.bl,th{{font-family:'Poppins',-apple-system,Segoe UI,Roboto,Arial,sans-serif}}
.wrap{{max-width:1320px;margin:0 auto;padding:24px}}
h1{{font-size:22px;margin:0 0 2px}}.sub{{color:#555;font-size:13px;margin-bottom:20px}}
.kpis{{display:grid;grid-template-columns:repeat(6,1fr);gap:12px;margin-bottom:22px}}
.kpi{{background:#fff;border:1px solid #ececec;border-radius:12px;padding:14px}}
.kpi .n{{font-size:26px;font-weight:700}}.kpi .l{{color:#555;font-size:11px;text-transform:uppercase;letter-spacing:.03em}}
.kpi.venc .n{{color:#7b241c}}.kpi.r .n{{color:#c0392b}}.kpi.y .n{{color:#9a7d0a}}.kpi.g .n{{color:#1e7e34}}.kpi.conf .n{{color:#1e7e34}}.kpi.aconf .n{{color:#9a7d0a}}
.grid{{display:grid;grid-template-columns:1fr 1fr;gap:14px;margin-bottom:22px}}
.card{{background:#fff;border:1px solid #ececec;border-radius:12px;padding:16px}}
.card h3{{margin:0 0 12px;font-size:14px}}
.bar{{display:flex;align-items:center;gap:8px;margin:6px 0;font-size:12.5px}}
.bl{{width:200px;flex:none;color:#333;white-space:nowrap;overflow:hidden;text-overflow:ellipsis}}
.bt{{height:14px;border-radius:7px;display:inline-block}}.bn{{font-weight:600;min-width:24px;text-align:right}}
.filters{{display:flex;gap:10px;flex-wrap:wrap;margin-bottom:12px}}
select,input{{padding:8px 10px;border:1px solid #ddd;border-radius:8px;font-size:13px;background:#fff;font-family:'Lora',serif}}
table{{width:100%;border-collapse:collapse;background:#fff;font-size:12.5px}}
th,td{{text-align:left;padding:9px 10px;border-bottom:1px solid #f0f0f0;vertical-align:top}}
th{{background:#f7f7f7;cursor:pointer;position:sticky;top:0;z-index:1}}
tr{{border-left:5px solid transparent}}
tr.venc{{border-left-color:#7b241c;background:#fbeeec}}tr.r{{border-left-color:#e74c3c;background:#fdecea}}
tr.y{{border-left-color:#e0a800;background:#fff8e6}}tr.g{{border-left-color:#28a745;background:#f1faf3}}tr.sem{{border-left-color:#ccc}}
.pill{{display:inline-block;padding:2px 8px;border-radius:20px;font-size:11px;font-weight:600;color:#fff}}
.muted{{color:#999}}a{{color:#1155cc}}
@media(max-width:1000px){{.kpis{{grid-template-columns:repeat(3,1fr)}}.grid{{grid-template-columns:1fr}}}}
</style></head><body><div class=wrap>
<h1>Painel de Publicações — DOE-SC / DETRAN</h1>
<div class=sub>{esc(sub)}Gerado em {HOJE.strftime('%d/%m/%Y')} · retrato do dia (não recarrega ao vivo)</div>
<div class=kpis>
<div class=kpi><div class=n>{total}</div><div class=l>Publicações</div></div>
<div class="kpi conf"><div class=n>{conf}</div><div class=l>Confirmadas</div></div>
<div class="kpi aconf"><div class=n>{aconf}</div><div class=l>A conferir</div></div>
<div class="kpi venc"><div class=n>{cnt['venc']}</div><div class=l>Vencidas (est.)</div></div>
<div class="kpi r"><div class=n>{cnt['r']}</div><div class=l>≤{ARGS_URGENTE} dias</div></div>
<div class=kpi><div class=n>{cnt['sem']}</div><div class=l>Sem prazo</div></div>
</div>
<div class=grid>
<div class=card><h3>Por tipo de decisão</h3>{bars(por_tipo)}</div>
<div class=card><h3>Por situação de prazo</h3>{bars_estado()}</div>
</div>
<div class=card>
<div class=filters>
<input id=q placeholder="Buscar cliente, processo, CNH, decisão…" style=flex:1>
<select id=fsb><option value="">Confirmada e a conferir</option><option value="CONFIRMADO">Só confirmadas</option><option value="A CONFERIR">Só a conferir</option></select>
<select id=ftipo><option value="">Todos os tipos</option>{opt(tipos)}</select>
<select id=festado><option value="">Toda situação</option><option value=venc>Vencida (est.)</option><option value=r>≤{ARGS_URGENTE} dias</option><option value=y>6–10 dias</option><option value=g>11+ dias</option><option value=sem>Sem prazo</option></select>
</div>
<div style="max-height:560px;overflow:auto"><table id=tb><thead><tr>
<th>Data final ▲</th><th>Dias</th><th>Publicação</th><th>Cliente</th><th>CNH</th><th>Processo</th><th>Tipo</th><th>Status busca</th><th>Status controle</th><th>Link</th>
</tr></thead><tbody id=body>{TBODY}</tbody></table></div>
<div id=vazio class=muted style="display:none;padding:14px">Nenhum item com os filtros atuais.</div>
</div>
<p class=muted style=font-size:11px>Semáforo por dias corridos até a data final estimada de recurso. Os prazos são apoio — confira sempre nos autos. Dados pessoais para uso interno (LGPD).</p>
</div>
<script>
const $=s=>document.querySelector(s);
function f(){{const q=$('#q').value.toLowerCase(),sb=$('#fsb').value,t=$('#ftipo').value,e=$('#festado').value;let vis=0;
 document.querySelectorAll('#body tr').forEach(tr=>{{
  let ok=(!q||tr.dataset.txt.includes(q))&&(!sb||tr.dataset.sb===sb)&&(!t||tr.dataset.tipo===t)&&(!e||tr.dataset.state===e);
  tr.style.display=ok?'':'none';if(ok)vis++;}});
 $('#vazio').style.display=vis?'none':'block';}}
['q','fsb','ftipo','festado'].forEach(id=>$('#'+id).addEventListener('input',f));
</script></body></html>"""
    open(DASH, "w", encoding="utf-8").write(HTML)
    versoes = os.path.join(saida, "Versões")
    os.makedirs(versoes, exist_ok=True)
    shutil.copy2(DASH, os.path.join(versoes, f"Dashboard_DOE_{HOJE.strftime('%Y-%m-%d')}.html"))
    return DASH


def main():
    args = parse_args()
    global ARGS_URGENTE
    ARGS_URGENTE = args.urgente

    regs, origem = carregar_registros(args)
    saida = args.saida or (os.path.dirname(origem) if os.path.isfile(origem) else origem) or "."
    saida = os.path.abspath(saida)
    os.makedirs(saida, exist_ok=True)

    if not regs:
        print(json.dumps({"ok": True, "total": 0, "aviso": "Nenhum registro para montar o relatório.", "origem": origem}, ensure_ascii=False))
        return

    # classifica e ordena
    for r in regs:
        df = parse_data(r["dataFinal"])
        r["_df"] = df
        st, dias = estado(df)
        r["_state"] = st
        r["_dias"] = dias
    regs.sort(key=lambda r: (r["_df"] or datetime.date.max, ORD[r["_state"]], 0 if r["statusBusca"] == "CONFIRMADO" else 1, r["cliente"]))

    xlsx, cnt = gerar_xlsx(regs, saida, args.titulo)
    dash = gerar_html(regs, saida, args.titulo, cnt)

    conf = sum(1 for r in regs if r["statusBusca"] == "CONFIRMADO")
    resumo = {
        "ok": True,
        "total": len(regs),
        "confirmados": conf,
        "aConferir": len(regs) - conf,
        "vencidos": cnt["venc"],
        "urgentes": cnt["venc"] + cnt["r"],
        "semPrazo": cnt["sem"],
        "planilha": xlsx,
        "dashboard": dash,
        "versoes": os.path.join(saida, "Versões"),
    }
    print("__RESUMO_RELATORIO__")
    print(json.dumps(resumo, ensure_ascii=False))


if __name__ == "__main__":
    main()
